from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from automation.core.task_contract import TaskResult  # pyright: ignore[reportMissingImports]
from .loader import load_data
from .formatter import normalize_rows
try:
    # Preferred (CLI / full system)
    from automation.utils.excel_writer import write_excel # pyright: ignore[reportMissingImports]
except ImportError:
    # Standalone fallback
    from .internal.excel_writer import write_excel

PARAM_SPEC = {
    "input_path": {
        "required": True,
        "description": "Path to ETL output JSON file"
    },
    "output_name": {
        "required": False,
        "default": "report.xlsx",
        "description": "Output Excel file name"
    },
    "exclude_errors": {
        "required": False,
        "default": False,
        "description": "Exclude rows with errors"
    },
    "columns": {
        "required": False,
        "description": "Pipe-separated columns to include (e.g. source|temperature|time)"
    },
    "output_mode": {
        "required": False,
        "default": "wide",
        "description": "wide, grouped, or sheets"
    },
    "mapping_file": {
    "required": False,
    "description": "Path to TOML extraction mapping"
    }
}


TASK_NAME = "excel_kpi"


def run(artifact_dir: Path, params: dict):

    mapping_file = params.get("mapping_file")

    mapping_config = None

    if mapping_file:
        from automation.utils.path_resolver import resolve_path # pyright: ignore[reportMissingImports]

        mapping_path = resolve_path(mapping_file, "mappings")

        import toml
        mapping_config = toml.load(mapping_path)

        mapping_config = {
            key.lower(): value
            for key, value in mapping_config.items()
        }

    try:
        input_path = Path(params["input_path"])
        output_name = params.get("output_name", "report.xlsx")

        data = load_data(input_path)

        rows = normalize_rows(
            data,
            exclude_errors=params.get("exclude_errors", False),
            columns=params.get("columns")
        )

        # ========================
        # CALCULATE METRICS (FIXED ORDER)
        # ========================
        total = len(rows)

        success_count = sum(1 for r in rows if r.get("status") == "SUCCESS")
        failure_count = sum(1 for r in rows if r.get("status") == "FAILURE")

        success_rate = (success_count / total * 100) if total > 0 else 0

        summary_data = [
            ["Metric", "Value"],
            ["Total", total],
            ["Success", success_count],
            ["Failure", failure_count],
            ["Success Rate", f"{success_rate:.2f}%"]
        ]

        output_mode = params.get("output_mode", "wide")

        output_file = artifact_dir / output_name

        if not rows:
            raise RuntimeError("No data available for Excel report")

        write_excel(
            output_file,
            rows,
            mode=output_mode,
            mapping_config=mapping_config
        )

        wb = load_workbook(output_file)

        if "Summary" in wb.sheetnames:
            del wb["Summary"]

        ws = wb.create_sheet(title="Summary")

        for row in summary_data:
            ws.append(row)

        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Auto width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(output_file)

        return TaskResult(
            success=True,
            message=f"Excel report generated: {output_file}",
            artifacts=[str(output_file)]
        )

    except Exception as e:
        return TaskResult(
            success=False,
            message=str(e)
        )