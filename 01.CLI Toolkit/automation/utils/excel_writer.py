from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from collections import defaultdict


# ========================
# HELPERS
# ========================

def _auto_size_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2


def _apply_status_colors(ws, status_col_index, start_row=2):
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=start_row):
        cell = row[status_col_index - 1]

        if cell.value == "SUCCESS":
            cell.fill = green
        elif cell.value == "FAILURE":
            cell.fill = red


def _safe_sheet_name(name):
    # Excel max = 31 chars, no special chars like /
    return str(name)[:31].replace("/", "_")


# ========================
# MAIN ROUTER
# ========================

def write_excel(output_path, data, mode="wide", mapping_config=None):

    valid_modes = {"wide", "grouped", "sheets"}

    if mode not in valid_modes:
        raise ValueError(f"Invalid output_mode: {mode}")

    if not data:
        raise ValueError("No data provided to write_excel")

    if mode == "grouped":
        return write_excel_grouped(output_path, data, mapping_config)

    elif mode == "sheets":
        return write_excel_sheets(output_path, data, mapping_config)

    # ========================
    # WIDE MODE
    # ========================

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    mapping_config = mapping_config or {}

    ordered_keys = ["source", "status"]

    for row in data:
        for k in row.keys():
            if k not in ordered_keys:
                ordered_keys.append(k)

    ws.append(ordered_keys)
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in data:
        ws.append([row.get(k) for k in ordered_keys])

    _auto_size_columns(ws)

    ws.auto_filter.ref = ws.dimensions

    status_col_index = ordered_keys.index("status") + 1
    _apply_status_colors(ws, status_col_index)

    wb.save(output_path)


# ========================
# GROUPED MODE
# ========================

def write_excel_grouped(output_path, data, mapping_config=None):

    mapping_config = mapping_config or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    grouped = defaultdict(list)

    for row in data:
        grouped[row["source"]].append(row)

    current_row = 1

    for source in sorted(grouped.keys()):
        rows = grouped[source]

        # ========================
        # BUILD KEYS PER SOURCE
        # ========================
        mapping = mapping_config.get(source.lower())

        if not mapping:
            for key, value in mapping_config.items():
                aliases = value.get("aliases", [])
                if source.lower() in [a.lower() for a in aliases]:
                    mapping = value
                    break

        mapping = mapping or {}

        ordered_keys = ["source", "status"] + [
            k for k in mapping.keys() if k != "aliases"
        ]

        # ========================
        # HEADER
        # ========================
        for col_idx, key in enumerate(ordered_keys, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=key)
            cell.font = Font(bold=True)

        header_row = current_row
        current_row += 1

        # ========================
        # DATA ROWS
        # ========================
        for row in rows:
            for col_idx, key in enumerate(ordered_keys, start=1):
                ws.cell(row=current_row, column=col_idx, value=row.get(key))
            current_row += 1

        # ========================
        # STATUS COLORS
        # ========================
        status_col_index = ordered_keys.index("status") + 1
        _apply_status_colors(ws, status_col_index, start_row=header_row + 1)

        # ========================
        # SPACING BETWEEN TABLES
        # ========================
        current_row += 1  # ONE empty row between groups

    _auto_size_columns(ws)

    wb.save(output_path)


# ========================
# SHEETS MODE
# ========================

def write_excel_sheets(output_path, data, mapping_config=None):

    mapping_config = mapping_config or {}

    wb = Workbook()
    wb.remove(wb.active)

    grouped = defaultdict(list)

    for row in data:
        grouped[row["source"]].append(row)

    for source in sorted(grouped.keys()):
        rows = grouped[source]

        sheet_name = _safe_sheet_name(source)
        ws = wb.create_sheet(title=sheet_name)

        mapping = mapping_config.get(source.lower())

        if not mapping:
            for key, value in mapping_config.items():
                aliases = value.get("aliases", [])
                if source.lower() in [a.lower() for a in aliases]:
                    mapping = value
                    break

        mapping = mapping or {}

        ordered_keys = ["source", "status"] + [
            k for k in mapping.keys() if k != "aliases"
        ]

        ws.append(ordered_keys)

        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"

        for row in rows:
            ws.append([row.get(k) for k in ordered_keys])

        _auto_size_columns(ws)

        ws.auto_filter.ref = ws.dimensions

        status_col_index = ordered_keys.index("status") + 1
        _apply_status_colors(ws, status_col_index)

    wb.save(output_path)